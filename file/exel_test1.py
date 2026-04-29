#엑셀 파일 다루기
#엑셀 확장자는 .xlsx 표 형태의 데이터 저장 파일이다.
#엑셀 다루기 위해 엑셀 라이브러리가 필요하다.  있으면 편함
#없으면 하나하나엑셀이라는 파일 인식 부터 코드 다 만들어야됨
#그래서 라이브러리를 설치한다. 파이썬의 라이브러리 설치는
#인터넷에서 찾아서 하는게 아니라 명령어로 작업한다.
#pip install openpyxl - 엑셀 라이브러리 설치

# from openpyxl import Workbook

# wb = Workbook()
# ws = wb.active

# ws["A1"] = "이름"
# ws["B1"] = "나이"
# ws["A2"] = "김유신" 
# ws["B2"] = "32"

# wb.save("file/test1.xlsx")

#엑셀 읽기
from openpyxl import load_workbook
wb = load_workbook("file/test1.xlsx")
ws = wb.active #



# print(ws["A2"].value)


# for row in ws.iter_rows():

    # for cell in row:
    #     print(cell.value)

#행단위로 읽어오기

# data_list ==[]

# for row in ws.iter_rows(values_only=True)
#     print( row)
#     data_list.append(list(row))

wb2 = load_workbook("file/test1.xlsx")

print(wb2.sheetnames) #엑셀 파일의 시트 목록 출력

ws = wb2["Sheet1"]

print( ws["a1"].value)